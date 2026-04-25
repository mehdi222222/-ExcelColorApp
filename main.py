import kivy
kivy.require("2.1.0")

from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.filechooser import FileChooserIconView
from kivy.uix.widget import Widget
from kivy.graphics import Color, Rectangle

import openpyxl


def excel_color_to_hex(cell):
    """Extract RGB hex from cell color."""
    try:
        fill = cell.fill
        if fill is None or fill.fgColor is None:
            return "#FFFFFF"
        fg = fill.fgColor

        if fg.type == "rgb":
            rgb = fg.rgb
            if rgb and len(rgb) == 8:
                return "#" + rgb[2:]
    except:
        pass
    return "#FFFFFF"


class ColorBox(Widget):
    def __init__(self, hex_color="#FFFFFF", **kwargs):
        super().__init__(**kwargs)
        self.hex = hex_color
        with self.canvas:
            Color(*self.hex_to_rgb(self.hex))
            self.rect = Rectangle(pos=self.pos, size=self.size)
        self.bind(pos=self.update_rect, size=self.update_rect)

    def update_color(self, hex_color):
        self.hex = hex_color
        r, g, b = self.hex_to_rgb(self.hex)
        self.canvas.clear()
        with self.canvas:
            Color(r, g, b)
            self.rect = Rectangle(pos=self.pos, size=self.size)

    def update_rect(self, *args):
        self.rect.pos = self.pos
        self.rect.size = self.size

    def hex_to_rgb(self, hx):
        hx = hx.lstrip('#')
        return tuple(int(hx[i:i+2], 16)/255 for i in (0,2,4))


class FileScreen(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)

        layout = BoxLayout(orientation="vertical")

        self.fc = FileChooserIconView(filters=["*.xlsx", "*.xlsm"])
        layout.add_widget(self.fc)

        btn = Button(text="باز کردن فایل", size_hint=(1, 0.15))
        btn.bind(on_release=self.load_excel)
        layout.add_widget(btn)

        self.add_widget(layout)

    def load_excel(self, instance):
        if len(self.fc.selection) == 0:
            return

        app = App.get_running_app()
        app.excel_path = self.fc.selection[0]

        self.manager.current = "viewer"


class Viewer(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)

        self.layout = BoxLayout(orientation="vertical", padding=10, spacing=10)

        self.info = Label(text="Cell: -", font_size=28)
        self.layout.add_widget(self.info)

        self.color_box = ColorBox("#FFFFFF", size_hint=(1, 0.5))
        self.layout.add_widget(self.color_box)

        btn_next = Button(text="Next Cell", size_hint=(1, 0.15))
        btn_next.bind(on_release=self.next_cell)
        self.layout.add_widget(btn_next)

        btn_reset = Button(text="Reset", size_hint=(1, 0.15))
        btn_reset.bind(on_release=self.reset_reader)
        self.layout.add_widget(btn_reset)

        self.add_widget(self.layout)

        self.ws = None
        self.max_row = 0
        self.max_col = 0
        self.row = 0
        self.col = 0

    def on_enter(self):
        self.load_excel()

    def load_excel(self):
        app = App.get_running_app()
        wb = openpyxl.load_workbook(app.excel_path)
        ws = wb.active

        self.ws = ws
        self.max_row = ws.max_row
        self.max_col = ws.max_column

        self.row = self.max_row
        self.col = self.max_col

        self.update_display()

    def next_cell(self, instance):
        if self.row < 1:
            return

        self.col -= 1

        if self.col < 1:
            self.row -= 1
            if self.row < 1:
                self.info.text = "پایان فایل"
                return
            self.col = self.max_col

        self.update_display()

    def reset_reader(self, instance):
        self.row = self.max_row
        self.col = self.max_col
        self.update_display()

    def update_display(self):
        try:
            cell = self.ws.cell(row=self.row, column=self.col)
            color = excel_color_to_hex(cell)

            cell_name = cell.coordinate
            self.info.text = f"Cell: {cell_name}"

            self.color_box.update_color(color)

        except Exception as e:
            self.info.text = f"Error: {e}"
            self.color_box.update_color("#FFFFFF")


class ExcelApp(App):
    excel_path = None

    def build(self):
        sm = ScreenManager()
        sm.add_widget(FileScreen(name="file"))
        sm.add_widget(Viewer(name="viewer"))
        return sm


if __name__ == '__main__':
    ExcelApp().run()
